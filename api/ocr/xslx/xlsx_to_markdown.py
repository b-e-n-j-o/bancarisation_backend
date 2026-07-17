"""
xlsx_to_markdown.py — Sérialisation d'un classeur Excel en markdown pour extraction LLM.

Équivalent de l'étape OCR du pipeline plan de gestion, mais sans OCR :
le xlsx est déjà du texte structuré, on le met en forme pour le LLM.

Principes :
- Cellules fusionnées dé-fusionnées (valeur du coin haut-gauche propagée sur la plage)
- Deux passes openpyxl : data_only=True (valeurs calculées) + passe formules
  → les lignes contenant des SUM/SUBTOTAL sont annotées [TOTAL?] pour aider le LLM
  à ne pas compter deux fois
- number_format inspecté pour annoter les montants (€) et pourcentages
- Flottants sales arrondis (7997.499999999 → 7997.5)
- Lignes/colonnes entièrement vides supprimées, coordonnées Excel conservées
  en tête de colonne pour la traçabilité (comme la page PDF dans les échéances)

Usage :
    python3 xlsx_to_markdown.py estimation.xlsx > estimation.md
    python3 xlsx_to_markdown.py estimation.xlsx --sheet "Planning"
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Formats numériques → suffixe d'annotation
CURRENCY_HINTS = ("€", "$", "£", '"EUR"', "#,##0.00 ", "_-* #")
PERCENT_HINT = "%"
TOTAL_FORMULA_RE = re.compile(
    r"(?:SUM|SUBTOTAL|SOMME)\s*\(\s*\$?([A-Z]{1,3})\$?(\d+)\s*:\s*\$?([A-Z]{1,3})\$?(\d+)",
    re.IGNORECASE,
)


def _is_vertical_sum(formula: str) -> bool:
    """Vrai si la formule contient une somme sur une plage multi-lignes
    (agrégation verticale = ligne de total), pas une somme horizontale
    type TTC = SUM(C5:D5)."""
    for m in TOTAL_FORMULA_RE.finditer(formula):
        col1, row1, col2, row2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        if col1 == col2 and row2 > row1:
            return True
    return False


def _clean_number(value: float) -> str:
    """Arrondit les flottants sales et supprime les .0 inutiles."""
    rounded = round(value, 2)
    if rounded == int(rounded):
        return str(int(rounded))
    return f"{rounded:g}"


def _format_value(value, number_format: str) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "VRAI" if value else "FAUX"
    if isinstance(value, (int, float)):
        text = _clean_number(float(value))
        nf = number_format or ""
        if any(h in nf for h in CURRENCY_HINTS):
            return f"{text} €"
        if PERCENT_HINT in nf:
            # openpyxl renvoie la fraction (0.2 → 20%)
            return f"{_clean_number(float(value) * 100)} %"
        return text
    # Texte : neutraliser retours à la ligne et pipes (casseraient la table md)
    return str(value).replace("\n", " / ").replace("|", "¦").strip()


class MergeInfo:
    """Métadonnées d'une cellule appartenant à une plage fusionnée."""

    __slots__ = ("value", "anchor", "min_row", "max_row", "min_col", "max_col")

    def __init__(self, value, anchor, min_row, max_row, min_col, max_col):
        self.value = value
        self.anchor = anchor  # (row, col) du coin haut-gauche
        self.min_row, self.max_row = min_row, max_row
        self.min_col, self.max_col = min_col, max_col

    @property
    def is_anchor_of(self):
        return self.anchor

    @property
    def spans_rows(self) -> bool:
        return self.max_row > self.min_row

    @property
    def spans_cols(self) -> bool:
        return self.max_col > self.min_col


def _merged_lookup(ws: Worksheet) -> dict[tuple[int, int], MergeInfo]:
    """Map (row, col) → MergeInfo pour toute cellule d'une plage fusionnée.

    On ne propage PAS aveuglément : le sens d'une fusion dépend de son contenu.
      - texte fusionné  = étiquette catégorielle du groupe → à propager
        (ex. B32:B34 = 'Pilotage et coordination' : les 3 lignes sont de cette famille)
      - nombre fusionné = quantité DU GROUPE, pas de chaque ligne → à ne pas propager
        (ex. E32:E34 = 2425 : les 3 prestations coûtent 2425 € ensemble, pas 2425 chacune)
    La distinction est appliquée dans sheet_to_markdown().
    """
    lookup: dict[tuple[int, int], MergeInfo] = {}
    for mr in ws.merged_cells.ranges:
        anchor = (mr.min_row, mr.min_col)
        value = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                lookup[(r, c)] = MergeInfo(
                    value, anchor, mr.min_row, mr.max_row, mr.min_col, mr.max_col
                )
    return lookup


def sheet_to_markdown(ws_values: Worksheet, ws_formulas: Worksheet) -> str:
    """Sérialise une feuille en table markdown unique, coordonnées incluses."""
    merged = _merged_lookup(ws_values)
    max_row, max_col = ws_values.max_row, ws_values.max_column

    # 1. Grille de textes formatés
    grid: list[list[str]] = []
    total_rows: set[int] = set()
    row_notes: dict[int, str] = {}
    for r in range(1, max_row + 1):
        row_cells: list[str] = []
        spans_down: set[tuple[int, int]] = set()   # fusions numériques ancrées ici
        spans_up: set[int] = set()                 # fusions numériques portées plus haut
        for c in range(1, max_col + 1):
            mi = merged.get((r, c))
            nf = ws_values.cell(r, c).number_format
            if mi is None:
                row_cells.append(_format_value(ws_values.cell(r, c).value, nf))
            elif not isinstance(mi.value, (int, float)) or isinstance(mi.value, bool):
                # Étiquette : la fusion signifie "ces cellules appartiennent à X"
                row_cells.append(_format_value(mi.value, nf))
            else:
                # Quantité : la valeur appartient à la plage entière, une seule fois
                nf_anchor = ws_values.cell(*mi.anchor).number_format
                if (r, c) == mi.anchor:
                    txt = _format_value(mi.value, nf_anchor)
                    if mi.spans_cols:
                        txt += (f" ⟨col. {get_column_letter(mi.min_col)}"
                                f"-{get_column_letter(mi.max_col)}⟩")
                    row_cells.append(txt)
                    if mi.spans_rows:
                        spans_down.add((mi.min_row, mi.max_row))
                else:
                    row_cells.append("")
                    if mi.spans_rows:
                        spans_up.add(mi.min_row)
            # Détection totaux via la passe formules
            fval = ws_formulas.cell(r, c).value
            if isinstance(fval, str) and _is_vertical_sum(fval):
                total_rows.add(r)
        grid.append(row_cells)

        notes = []
        for lo, hi in sorted(spans_down):
            notes.append(f"montants communs aux lignes {lo}-{hi}")
        for anchor_row in sorted(spans_up):
            notes.append(f"montants portés par la ligne {anchor_row}")
        if notes:
            row_notes[r] = "; ".join(notes)

    # 2. Colonnes entièrement vides → supprimées (index conservés pour l'en-tête)
    kept_cols = [c for c in range(max_col) if any(row[c] for row in grid)]
    if not kept_cols:
        return "_(feuille vide)_\n"

    # 3. Lignes vides → supprimées ; numéro de ligne Excel conservé
    lines: list[str] = []
    header = ["ligne"] + [get_column_letter(c + 1) for c in kept_cols]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("|" + "---|" * len(header))
    for r_idx, row in enumerate(grid, start=1):
        cells = [row[c] for c in kept_cols]
        if not any(cells) and r_idx not in row_notes:
            continue
        marker = str(r_idx)
        if r_idx in total_rows:
            marker += " [TOTAL?]"
        if r_idx in row_notes:
            marker += f" ⟨{row_notes[r_idx]}⟩"
        lines.append("| " + marker + " | " + " | ".join(cells) + " |")
    return "\n".join(lines) + "\n"


def workbook_to_markdown(path: str | Path, only_sheet: str | None = None) -> str:
    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path)  # formules brutes

    parts: list[str] = [f"# Classeur : {Path(path).name}\n"]
    for name in wb_values.sheetnames:
        if only_sheet and name != only_sheet:
            continue
        ws_v, ws_f = wb_values[name], wb_formulas[name]
        parts.append(f"\n## Feuille : {name}\n")
        parts.append(sheet_to_markdown(ws_v, ws_f))
    return "\n".join(parts)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", help="Chemin du fichier .xlsx")
    parser.add_argument("--sheet", help="Ne sérialiser qu'une feuille", default=None)
    args = parser.parse_args()
    sys.stdout.write(workbook_to_markdown(args.xlsx, args.sheet))